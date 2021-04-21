from matplotlib import pyplot
from openpyxl import load_workbook

#dir.load_workbook()
#!!!Нужно как-то добавить путь чтобы не писать полный путь.
wb=load_workbook('C:\\Users\\Kalinichev\\Documents\\VisualStudioCode_local\\dakalinichev_p4ne\\p4ne\Lab1.2\\data_analysis_lab.xlsx')# Загрузить таблицу Excel из файла в переменную wb
sheet = wb['Data'] # Загрузить лист с именем "Data" в переменную sheet #type(sheet)
sheet.values['A'][1:] # Получить содержимое колонки A в виде списка

def getvalue(x):
    return x.value


list(map(getvalue, sheet['A'][1:]))
map(getvalue, sheet['A'][1:]) # Преобразовать содержимое колонки A в список, содержащий только значения (без форматирования и т. п.)
List1=map(getvalue, sheet['A'][1:])
#dir(list_y)
#dir(map)    type(list_x)
list_x=list(map(getvalue, sheet['A'][1:]))
list_y2=list(map(getvalue, sheet['D'][1:]))
list_y=list(map(getvalue, sheet['C'][1:]))
# Построить график по точкам, в первом списке значения по оси X, во втором — значения по оси Y 
"""
https://matplotlib.org/stable/api/_as_gen/matplotlib.pyplot.plot.html
Line Styles
    'b'    # blue markers with default shape
    'or'   # red circles
    '-g'   # green solid line
    '--'   # dashed line with default color
    '^k:'  # black triangle_up markers connected by a dotted line
Color
    'b'	blue
    'g'	green
    'r'	red
    'c'	cyan
    'm'	magenta
    'y'	yellow
    'k'	black
    'w'	white

"""

pyplot.plot(list_x, list_y, linewidth=2,label="Относит. температура",color="c") 
pyplot.plot(list_x, list_y2, linewidth=2,label="Активность",linestyle='dashed',color="m")

# Украшаем график - добавляем надписи
pyplot.xlabel('Годы')
pyplot.ylabel('Температура/Активность Солнца')
pyplot.legend(loc='upper left') # Отображаем легенду
pyplot.show() # показать график


